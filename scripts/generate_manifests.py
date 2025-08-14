import yaml
import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def load_yaml(file_path):
    with open(file_path, 'r') as f:
        return yaml.safe_load(f)

def get_module_name(yaml_path):
    """Get the correct module name based on the YAML file name"""
    name_map = {
        'demographics.yaml': 'Demographics',
        'diagnosis.yaml': 'Diagnosis',
        'exposure.yaml': 'Exposure',
        'family_history.yaml': 'Family History',
        'followup.yaml': 'Followup',
        'molecular.yaml': 'Molecular Test',
        'therapy.yaml': 'Therapy',
        'vital_status.yaml': 'Vital Status'
    }
    return name_map.get(os.path.basename(yaml_path), os.path.splitext(os.path.basename(yaml_path))[0].capitalize())

def get_enum_values(schema, enum_name):
    """Get permissible values for an enum"""
    if 'enums' in schema and enum_name in schema['enums']:
        return list(schema['enums'][enum_name].get('permissible_values', {}).keys())
    return []

def create_module_sheet(writer, schema, module_name):
    """Create a sheet for a module with its attributes"""
    if 'classes' not in schema:
        return
    
    # Get the main class (first one in the schema)
    class_name = next(iter(schema['classes']))
    class_def = schema['classes'][class_name]
    
    # Create DataFrame with attribute names as columns
    attributes = list(class_def.get('attributes', {}).keys())
    df = pd.DataFrame(columns=attributes)
    
    # Write to Excel
    df.to_excel(writer, sheet_name=module_name, index=False)
    
    # Get the worksheet
    worksheet = writer.sheets[module_name]
    
    # Add dropdowns for enum fields
    for col_num, attr_name in enumerate(attributes, 1):
        attr_def = class_def['attributes'][attr_name]
        range_type = attr_def.get('range', '')
        
        # Check if it's an enum
        if range_type.endswith('Enum'):
            enum_values = get_enum_values(schema, range_type)
            
            # Only add dropdown if enum has reasonable number of values
            if enum_values and len(enum_values) <= 100:
                col_letter = get_column_letter(col_num)
                dv = DataValidation(type="list", formula1=f'"{",".join(enum_values)}"', allow_blank=True)
                dv.add(f'{col_letter}2:{col_letter}1048576')  # Apply to all rows
                worksheet.add_data_validation(dv)
            elif enum_values:
                # Add a note about checking documentation
                worksheet.cell(row=1, column=col_num).comment = f"Many permissible values. See documentation for details."

def main():
    # Set up paths
    base_dir = Path(__file__).parent
    yaml_dir = base_dir / "modules" / "Clinical" / "domains"
    output_file = base_dir / "htan_clinical_manifest.xlsx"
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Process each YAML file
        for yaml_file in yaml_dir.glob("*.yaml"):
            if yaml_file.name not in ["clinical.yaml", "config.yaml"]:
                try:
                    schema = load_yaml(yaml_file)
                    module_name = get_module_name(yaml_file)
                    create_module_sheet(writer, schema, module_name)
                    print(f"Generated sheet for {module_name}")
                except Exception as e:
                    print(f"Error processing {yaml_file.name}: {str(e)}")
    
    print(f"Manifest generated: {output_file}")

if __name__ == "__main__":
    main() 